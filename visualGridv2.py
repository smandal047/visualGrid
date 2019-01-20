from copy import deepcopy
from openpyxl.styles import Border, Side, PatternFill, alignment


class VisualGridMaker(object):

    def __init__(self, info, worksheet, win_lines):
        self.ws = worksheet
        self.info = info
        self.win_lines = win_lines
        self.xcel_info = {
            'cell_start_row': 3,
            'cell_start_column': 1,
            'cell_end_row': 2 + info['view_size'],
            'cell_end_column': info['num_reels'],
            'shift': 2 + info['view_size']
        }
        self.grid_division = 3

    def setting_column_width(self):
        # setting the width of the columns, usually min is 1 and max is reel length
        for col in self.ws.iter_cols(min_col=6, max_col=25):
            col_name = col[0].column
            self.ws.column_dimensions[col_name].width = 3

    def write_to_excel_n_shift(self, grid, grid_start_offset=0):

        sheet_row = 0

        if grid_start_offset == 0:
            grid_start_offset = self.info['win_lines'] - len(grid)
        else:
            grid_start_offset = grid_start_offset - len(grid)

        for _index, _grid in enumerate(grid):
            sheet_row += 2

            # merging the grid no cell
            # self.ws.merge_cells("{}:{}".format(self.ws.cell(row=sheet_row, column=1).coordinate,
            #                                    self.ws.cell(row=sheet_row, column=5).coordinate))

            grid_no = grid_start_offset + _index + 1

            self.ws.cell(row=sheet_row, column=1).value = "Line_" + str(grid_no)
            self.ws.cell(row=sheet_row, column=1).alignment = alignment.Alignment('center')
            for row, _line in enumerate(_grid):
                sheet_row += 1
                for col, reel_cell in enumerate(_line):
                    self.ws.cell(row=sheet_row, column=col + 1).value = reel_cell

        # inserting columns
        self.ws.insert_cols(idx=1, amount=self.info['num_reels'] + 1)

        return grid_start_offset

    @staticmethod
    def set_border(worksheet, cell_range):
        rows = worksheet[cell_range]
        side_thick = Side(border_style='thick', color="FF000000")
        side_thin = Side(border_style='thin', color="FF000000")

        rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
        max_y = len(rows) - 1  # index of the last row
        for pos_y, cells in enumerate(rows):
            max_x = len(cells) - 1  # index of the last cell
            for pos_x, cell in enumerate(cells):
                border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom,
                )

                # applying the thin borders for the symbol grid
                border.bottom = border.top = border.right = border.left = side_thin
                cell.border = border

                # applying the thick borders around the grid
                if pos_x == 0:
                    border.left = side_thick
                if pos_x == max_x:
                    border.right = side_thick
                if pos_y == 0:
                    border.top = side_thick
                if pos_y == max_y:
                    border.bottom = side_thick

                # set new border only if it's one of the edge cells
                # if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border

    def apply_border(self, grid):
        shift = 0
        for _grid in range(len(grid)):
            # cell start and end need to be given
            cell_start = self.ws.cell(row=self.xcel_info['cell_start_row'] + shift,
                                      column=self.xcel_info['cell_start_column'])
            cell_end = self.ws.cell(row=self.xcel_info['cell_end_row'] + shift,
                                    column=self.xcel_info['cell_end_column'])

            shift += self.xcel_info['shift']

            # print(cell_start.coordinate, cell_end.coordinate)
            self.set_border(self.ws, "{}:{}".format(cell_start.coordinate, cell_end.coordinate))

    def grid_maker(self):
        # creating a sample list like below as per the symbol grid:
        # sample = [['\t', '\t', '\t', '\t', '\t'],
        #           ['\t', '\t', '\t', '\t', '\t'],
        #           ['\t', '\t', '\t', '\t', '\t']]
        sample = []
        for size in range(self.info['view_size']):
            _line = []
            for reel in range(self.info['num_reels']):
                _line.append("\t")
            sample.append(_line)

        # print(sample)

        # making the required grid from sample list
        grids_for_xcel = []
        for _line in self.win_lines:
            copy_sample = deepcopy(sample)
            for _col, _row in enumerate(_line):
                copy_sample[_row][_col] = 'X'
            grids_for_xcel.append(copy_sample)

        # print(grids_for_xcel)
        division = len(grids_for_xcel) // self.grid_division
        remainder = len(grids_for_xcel) % self.grid_division

        if remainder == 1:
            # adding 1 extra grid to list 1
            grid1_for_excel = grids_for_xcel[:division + 1]
            grid2_for_excel = grids_for_xcel[division + 1: (2 * division) + 1]
            grid3_for_excel = grids_for_xcel[(2 * division) + 1:]

        elif remainder == 2:
            # add 1 extra grid to both list 1 and list 2
            grid1_for_excel = grids_for_xcel[:division + 1]
            grid2_for_excel = grids_for_xcel[division + 1: (2 * division) + 2]
            grid3_for_excel = grids_for_xcel[(2 * division) + 2:]
        else:
            # nothing to add in any list
            grid1_for_excel = grids_for_xcel[:division]
            grid2_for_excel = grids_for_xcel[division:2 * division]
            grid3_for_excel = grids_for_xcel[2 * division:]

        # print(grid1_for_excel)
        # print(grid2_for_excel)
        # print(grid3_for_excel)

        # applying the border to the grid
        self.apply_border(grid3_for_excel)
        # writing the grid to worksheet
        grid_no = self.write_to_excel_n_shift(grid3_for_excel)

        self.apply_border(grid2_for_excel)
        grid_no = self.write_to_excel_n_shift(grid2_for_excel, grid_no)

        self.apply_border(grid1_for_excel)
        self.write_to_excel_n_shift(grid1_for_excel, grid_no)

        self.setting_column_width()

        # applying the fill for every X it finds
        for row in self.ws.iter_rows():
            # print(row[0])
            for cell in row:
                if cell.value == "X":
                    # apply color fill
                    red_fill = PatternFill(start_color='FFFF00',
                                           end_color='FFFF00',
                                           fill_type='solid')

                    self.ws[cell.coordinate].fill = red_fill